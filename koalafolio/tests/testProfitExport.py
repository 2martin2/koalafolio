# -*- coding: utf-8 -*-
"""
Created on Thu Jan 08 15:11:38 2019

@author: Martin
"""
import datetime
import koalafolio.PcpCore.settings as settings
import koalafolio.PcpCore.core as core
import koalafolio.Import.TradeImporter as importer
from export.profitExport import createProfitExcel

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# %% test

font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')
border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                                color='FF000000')
                )
alignment = Alignment(horizontal='general',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)
number_format = 'General'
protection = Protection(locked=True,
                        hidden=False)

# for cell in ws["2:2"]:
#    cell.font = red_font


# %% import data
mypath = r'D:\workspace\python\PyCryptoPortfolio\importdata'
# mypath = r'D:\workspace\spyder\PyCryptoPortfolio\data\kaj'
files, content = importer.loadTrades(mypath)

settings.mySettings.setPath('D:\workspace\python\PyCryptoPortfolio\Data')

headings = [frame.columns.tolist() for frame in content]
# fileheadings = []
# for i in range(len(files)):
#    fileheadings.append([files[i],headings[i]])


# %% convert data
# tradeList, matches = importer.convertTrades(models.IMPORT_MODEL_LIST, content)
tradeList = core.TradeList()
tradeList.fromCsv(r'D:\workspace\python\PyCryptoPortfolio\Data\Trades.csv')
# tradeList.updateValues()
tradeFrame = tradeList.toDataFrameComplete()

coinList = core.CoinList()
coinList.addTrades(tradeList)
coinList.updateCurrentValues()

coinFrame = coinList.toDataFrame()
coinDict = coinFrame.to_dict('index')

coinIndex = 3

# tradeMatcher = coinList.coins[coinIndex].tradeMatcher
# matchedFrame = coinList.coins[coinIndex].tradeMatcher.toDataFrame('EUR')

minDate = datetime.date(2017, 1, 1)
maxDate = datetime.date(2017, 12, 31)

createProfitExcel(coinList, 'test1', minDate, maxDate)

minDate = datetime.date(2018, 1, 1)
maxDate = datetime.date(2018, 12, 31)

createProfitExcel(coinList, 'test2', minDate, maxDate)