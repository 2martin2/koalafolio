# -*- coding: utf-8 -*-
"""
Created on Fri Aug 24 09:44:38 2018

@author: Martin
"""
import openpyxl
import openpyxl.styles.borders as xlborders
import koalafolio.PcpCore.settings as settings
import pytz
from dateutil.relativedelta import relativedelta
import koalafolio.PcpCore.core as core
import re
import koalafolio.gui.QLogger as logger

localLogger = logger.globalLogger

from openpyxl.styles import PatternFill, Alignment, Font


# colorcodes
BLUECOLORCODE = 'FF61D2FF'
GREENCOLORCODE = 'FF6DE992'
YELLOWCOLORCODE = 'FFFFFF52'
PURPLECOLORCODE = 'FFE057FF'
GRAYCOLORCODE = 'FFC8C8C8'

thin_border = xlborders.Border(left=xlborders.Side(style='thin', color=GRAYCOLORCODE),
                               right=xlborders.Side(style='thin', color=GRAYCOLORCODE),
                               top=xlborders.Side(style='thin', color=GRAYCOLORCODE),
                               bottom=xlborders.Side(style='thin', color=GRAYCOLORCODE),
                               vertical=xlborders.Side(style='thin', color=GRAYCOLORCODE),
                               horizontal=xlborders.Side(style='thin', color=GRAYCOLORCODE))


# %% create excelfile
def createProfitExcel(coinList, path, minDate, maxDate, currency='EUR', taxyearlimit=1,
                      includeTaxFreeTrades = True, lang="en", translator=None):

    if translator:
        def trans(text):
            return translator.translate(text, lang)

    wb = openpyxl.Workbook(write_only=False)
    wb.remove(wb.active)

    tradeprofitSumRows = []
    feeSumRows = []
    rewardSumRows = []
    tradeprofitSumColumns = []
    feeSumColumns = []
    rewardSumColumns = []

    # %% profit sheets
    profitSumColumn = 'O'
    profitSumRows = []
    profitSumColumns = []
    for coinWallets in coinList.coins:
        # check if coin is report Currency
        if not coinWallets.isReportCurrency():
            # check if there are sells in the given timeframe
            validSellFound = False
            for coin in coinWallets.wallets:
                for sell in coin.tradeMatcher.sellsMatched:
                    if sell.date >= minDate and sell.date <= maxDate:
                        validSellFound = True
                        break
                if validSellFound:
                    walletname = coin.getWalletName()
                    wsname = re.sub('[^A-Za-z0-9_]+', '', walletname)
                    ws = wb.create_sheet(wsname)

                    # write top header
                    # cell = WriteOnlyCell(ws)
                    if translator:
                        ws.append(['', '', '', trans('Buy'), '', '', '', '', trans('Sell'), '', '', '', '', trans('Profit'), ''])
                    else:
                        ws.append(['', '', '', 'Ankauf', '', '', '', '', 'Verkauf', '', '', '', '', 'Profit', ''])
                    ws.merge_cells('A1:B1')
                    ws.merge_cells('D1:G1')
                    ws.merge_cells('I1:L1')
                    ws.merge_cells('N1:O1')
                    headingFont = Font(size=12, bold=True)
                    headingAlignment = Alignment(horizontal='center',
                                                 vertical='center')
                    headings = [ws['A1'], ws['D1'], ws['I1'], ws['N1']]
                    for heading in headings:
                        heading.font = headingFont
                        heading.alignment = headingAlignment

                    # blue, green, yellow, purple
                    headingColors = [BLUECOLORCODE, GREENCOLORCODE, YELLOWCOLORCODE, PURPLECOLORCODE]
                    for i in range(len(headings)):
                        headings[i].fill = PatternFill(fill_type='solid',
                                                       start_color=headingColors[i],
                                                       end_color=headingColors[i])

                    # empty row
                    ws.append([])

                    # write sub header
                    if translator:
                        ws.append(
                            [trans('Coin') + '/' + trans('Wallet'), '', '', trans('Date'), trans('Amount'), trans('Price'), trans('Value'), '', trans('Date'),
                             trans('Amount'), trans('Price'), trans('Value'), '', trans('Profit'), trans('tax relevant')])
                        ws.append(['', '', '', '', trans('in') + ' ' + trans('pc'),
                                   trans('in') + ' ' + currency + '/' + trans('pc'),
                                   trans('in') + ' ' + currency, '', '', trans('in') + ' ' + trans('pc'),
                                   trans('in') + ' ' + currency + '/' + trans('pc'), trans('in') + ' ' + currency, '',
                                   trans('in') + ' ' + currency, trans('in') + ' ' + currency])
                    else:
                        ws.append(['Coin/Wallet', '', '', 'Datum', 'Anzahl', 'Preis', 'Wert', '', 'Datum', 'Anzahl', 'Preis', 'Wert', '',
                                   'Gewinn', 'zu versteuern'])
                        ws.append(['', '', '', '', 'in Stk', 'in ' + currency + '/Stk', 'in ' + currency, '', '', 'in Stk',
                                   'in ' + currency + '/Stk', 'in ' + currency, '', 'in ' + currency, 'in ' + currency])

                    # walletname
                    ws.append([wsname, ''])

                    firstProfitRow = ws.max_row + 1
                    # write data
                    rowcount = 0
                    for irow in range(len(coin.tradeMatcher.profitMatched)):
                        sell = coin.tradeMatcher.sellsMatched[irow]
                        # check amount of sell not zero (can happen for very small fees)
                        if sell.amount > 0:
                            # check date of sell
                            if sell.date >= minDate and sell.date <= maxDate:
                                buy = coin.tradeMatcher.buysMatched[irow]
                                profit = coin.tradeMatcher.profitMatched[irow]
                                # if taxyearlimit is given # if limit is relevant
                                if taxyearlimit and ((sell.date - relativedelta(years=taxyearlimit)) > buy.date):  # if taxyearlimit is given
                                    taxProfit = 0
                                    if includeTaxFreeTrades:
                                        rowcount += 1
                                        ws.append(
                                            ['', '', '', buy.date, buy.amount, buy.getPrice()[currency],
                                             buy.value[currency], '', sell.date, sell.amount, sell.getPrice()[currency],
                                             sell.value[currency], '', round(profit[currency], 3), round(taxProfit, 3)])
                                else:
                                    rowcount += 1
                                    taxProfit = profit[currency]
                                    ws.append(['', '', '', buy.date, buy.amount, buy.getPrice()[currency],
                                               buy.value[currency], '', sell.date, sell.amount, sell.getPrice()[currency],
                                               sell.value[currency], '', round(profit[currency], 3), round(taxProfit, 3)])
                    if rowcount == 0:  # no trades added:
                        wb.remove(ws)
                    else:
                        profitSumRows.append(ws.max_row + 2)
                        tradeprofitSumRows.append(profitSumRows[-1])
                        profitSumColumns.append(15)
                        tradeprofitSumColumns.append(profitSumColumns[-1])
                        #                ws['M' + str(profitSumRows[-1])] = 'Summe'
                        ws[profitSumColumn + str(profitSumRows[-1])] = '=ROUNDDOWN(SUM(' + profitSumColumn + str(
                            firstProfitRow) + ':' + profitSumColumn + str(profitSumRows[-1] - 2) + '),2)'

                        # page setup
                        if translator:
                            pageSetup(ws, dateCols=['D', 'I'], gapCols=['C', 'H', 'M', 'P'],
                                      lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                                      setWidthCols=['A', 'B', 'O'], setWidthValue=[15, 3, len(trans('tax relevant'))-1],
                                      trans=trans)
                        else:
                            pageSetup(ws, dateCols=['D', 'I'], gapCols=['C', 'H', 'M', 'P'],
                                      lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                                      setWidthCols=['A', 'B', 'O'], setWidthValue=[15, 3, 11])


    # %% fee sheets
    feeSumColumn = 'G'
    for coin in coinList.coins:
        fees = coin.getFees()
        feeSum = core.CoinValue()
        for fee in fees:
            if fee.date.date() >= minDate and fee.date.date() <= maxDate:
                feeSum.add(fee.value)
        if feeSum[currency] != 0:  # if fees have been paid
            wsname = re.sub('[^A-Za-z0-9]+', '', coin.coinname)
            ws = wb.create_sheet(wsname + '_fees')

            # write top header
            # cell = WriteOnlyCell(ws)
            if translator:
                ws.append(['', '', '', trans('Fees'), '', '', '', ''])
            else:
                ws.append(['', '', '', 'Gebühren', '', '', '', ''])
            ws.merge_cells('A1:B1')
            ws.merge_cells('D1:G1')
            headingFont = Font(size=12, bold=True)
            headingAlignment = Alignment(horizontal='center',
                                         vertical='center')
            headings = [ws['A1'], ws['D1']]
            for heading in headings:
                heading.font = headingFont
                heading.alignment = headingAlignment

            # blue, purple
            headingColors = [BLUECOLORCODE, PURPLECOLORCODE]
            for i in range(len(headings)):
                headings[i].fill = PatternFill(fill_type='solid',
                                               start_color=headingColors[i],
                                               end_color=headingColors[i])

            # empty row
            ws.append([])

            # write sub header
            if translator:
                ws.append(
                    [trans('Coin'), '', '', trans('Date'), trans('Fee'), '', trans('Fee'), ''])
                ws.append(['', '', '', '', trans('in') + ' ' + trans('pc'), '', trans('in') + ' ' + currency, ''])
            else:
                ws.append(
                    ['Coin', '', '', 'Datum', 'Gebühr', '', 'Gebühr', ''])
                ws.append(['', '', '', '', 'in Stk', '', 'in ' + currency, ''])

            # coinname
            ws.append([wsname, ''])

            firstProfitRow = ws.max_row + 1
            # write data
            fees.sort(key=lambda x: x.date, reverse=False)
            for fee in fees:
                # check amount of fee not zero
                if fee.amount < 0:
                    # check date of sell
                    if fee.date.date() >= minDate and fee.date.date() <= maxDate:
                        feedate = fee.date.astimezone(pytz.utc).replace(tzinfo=None).date()
                        ws.append(['', '', '', feedate, fee.amount, '', round(fee.value[currency], 3), ''])

            profitSumRows.append(ws.max_row + 2)
            feeSumRows.append(profitSumRows[-1])
            profitSumColumns.append(7)
            feeSumColumns.append(profitSumColumns[-1])
            #                ws['M' + str(profitSumRows[-1])] = 'Summe'
            ws[feeSumColumn + str(profitSumRows[-1])] = '=ROUNDDOWN(SUM(' + feeSumColumn + str(
                firstProfitRow) + ':' + feeSumColumn + str(profitSumRows[-1] - 2) + '),2)'

            # page setup
            # page setup
            if translator:
                pageSetup(ws, dateCols=['D'], gapCols=['C', 'H'],
                          lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                          setWidthCols=['A', 'B', 'F'], setWidthValue=[13, 3, 3], trans=trans)
            else:
                pageSetup(ws, dateCols=['D'], gapCols=['C', 'H'],
                          lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                          setWidthCols=['A', 'B', 'F'], setWidthValue=[13, 3, 3])


    # %% reward sheets
    rewardSumColumn = 'G'
    rewardSumRows = []
    for coin in coinList.coins:
        rewards = coin.getRewards()
        rewardSum = core.CoinValue()
        for reward in rewards:
            if reward.date.date() >= minDate and reward.date.date() <= maxDate:
                rewardSum.add(reward.value)
        if rewardSum[currency] != 0:  # if rewards have been received
            wsname = re.sub('[^A-Za-z0-9]+', '', coin.coinname)
            ws = wb.create_sheet(wsname + '_rewards')

            # write top header
            # cell = WriteOnlyCell(ws)
            if translator:
                ws.append(['', '', '', trans('Rewards'), '', '', '', ''])
            else:
                ws.append(['', '', '', 'Gebühren', '', '', '', ''])
            ws.merge_cells('A1:B1')
            ws.merge_cells('D1:G1')
            headingFont = Font(size=12, bold=True)
            headingAlignment = Alignment(horizontal='center',
                                         vertical='center')
            headings = [ws['A1'], ws['D1']]
            for heading in headings:
                heading.font = headingFont
                heading.alignment = headingAlignment

            # blue, purple
            headingColors = [BLUECOLORCODE, PURPLECOLORCODE]
            for i in range(len(headings)):
                headings[i].fill = PatternFill(fill_type='solid',
                                               start_color=headingColors[i],
                                               end_color=headingColors[i])

            # empty row
            ws.append([])

            # write sub header
            if translator:
                ws.append(
                    [trans('Coin'), '', '', trans('Date'), trans('Reward'), '', trans('Reward'), ''])
                ws.append(['', '', '', '', trans('in') + ' ' + trans('pc'), '', trans('in') + ' ' + currency, ''])
            else:
                ws.append(
                    ['Coin', '', '', 'Datum', 'Gebühr', '', 'Gebühr', ''])
                ws.append(['', '', '', '', 'in Stk', '', 'in ' + currency, ''])

            # coinname
            ws.append([wsname, ''])

            firstProfitRow = ws.max_row + 1
            # write data
            for reward in coin.getRewards():
                # check date of sell
                if reward.date.date() >= minDate and reward.date.date() <= maxDate:
                    rewarddate = reward.date.astimezone(pytz.utc).replace(tzinfo=None).date()
                    ws.append(['', '', '', rewarddate, reward.amount, '', round(reward.value[currency], 3), ''])

            profitSumRows.append(ws.max_row + 2)
            rewardSumRows.append(profitSumRows[-1])
            profitSumColumns.append(7)
            rewardSumColumns.append(profitSumColumns[-1])
            #                ws['M' + str(profitSumRows[-1])] = 'Summe'
            ws[rewardSumColumn + str(profitSumRows[-1])] = '=ROUNDDOWN(SUM(' + rewardSumColumn + str(
                firstProfitRow) + ':' + rewardSumColumn + str(profitSumRows[-1] - 2) + '),2)'

            # page setup
            if translator:
                pageSetup(ws, dateCols=['D'], gapCols=['C', 'H'],
                          lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                          setWidthCols=['A', 'B', 'F'], setWidthValue=[13, 3, 3], trans=trans)
            else:
                pageSetup(ws, dateCols=['D'], gapCols=['C', 'H'],
                          lastRow=profitSumRows[-1], lastCol=profitSumColumns[-1]+1,
                          setWidthCols=['A', 'B', 'F'], setWidthValue=[13, 3, 3])


    # %% overview sheet
    ws = wb.create_sheet('Overview', 0)

    # write top header
    if translator:
        ws.append(['', '', '', trans('Trades'), '', '', trans('Fees'), '', '', trans('Rewards'), '', ''])
    else:
        ws.append(['', '', '', 'Trades', '', '', 'Fees', '', '', 'Rewards', '', ''])
    ws.merge_cells('A1:B1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('J1:K1')
    headingFont = Font(size=12, bold=True)
    headingAlignment = Alignment(horizontal='center',
                                 vertical='center')

    # blue, green, yellow, purple
    headings = [ws['A1'], ws['D1'], ws['G1'], ws['J1']]
    for heading in headings:
        heading.font = headingFont
        heading.alignment = headingAlignment

    headingColors = [BLUECOLORCODE, GREENCOLORCODE, YELLOWCOLORCODE, PURPLECOLORCODE]
    for i in range(len(headings)):
        headings[i].fill = PatternFill(fill_type='solid',
                                       start_color=headingColors[i],
                                       end_color=headingColors[i])

    # empty row
    ws.append([])

    # write sub header
    if translator:
        ws.append(['', '', '', trans('Coin') + '/' + trans('Wallet'), trans('Profit'), '',
                   trans('Coin'), trans('Fee'), '', trans('Coin'), trans('Reward'), ''])
        ws.append(['', '', '', '', trans('in') + ' ' + currency, '',
                   '', trans('in') + ' ' + currency, '',
                   '', trans('in') + ' ' + currency, ''])
        ws.append([trans('Timeframe'), str(minDate) + ' : ' + str(maxDate)])
    else:
        ws.append(['', '', '', 'Coin/Wallet', 'Gewinn', '',
                   'Coin', 'Gebühren', '', 'Coin', 'Rewards', ''])
        ws.append(['', '', '', '', 'in ' + currency, '',
                   '', 'in ' + currency, '', '', 'in ' + currency, ''])
        ws.append(['Zeitraum', str(minDate) + ' : ' + str(maxDate)])

    firstProfitRow = ws.max_row + 1

    # write data
    sheets = wb.sheetnames[1:]
    # count number of different sheet types
    numTradeSheets = len(tradeprofitSumRows)
    numFeeSheets = len(feeSumRows)
    numRewardSheets = len(rewardSumRows)
    # for isheet in range(len(sheets)):
    for irow in range(max(numTradeSheets, numFeeSheets, numRewardSheets)):
        # claculate sheetindex for all sheet types
        iTradeSheet = irow
        iFeeSheet = irow + numTradeSheets
        iRewardSheet = iFeeSheet + numFeeSheets

        if irow < numTradeSheets:
            tradeProfitSumRef = '=INDIRECT(\"' + sheets[iTradeSheet] + '!\"&ADDRESS(' + str(tradeprofitSumRows[irow]) + ',' + str(tradeprofitSumColumns[irow]) + ',4))'
            tradeProfitSheet = sheets[iTradeSheet]
        else:
            tradeProfitSumRef = ''
            tradeProfitSheet = ''
        if irow < numFeeSheets:
            feeSumRef =  '=INDIRECT(\"' + sheets[iFeeSheet] + '!\"&ADDRESS(' + str(feeSumRows[irow]) + ',' + str(feeSumColumns[irow]) + ',4))'
            feeSheet = sheets[iFeeSheet]
        else:
            feeSumRef = ''
            feeSheet = ''
        if irow < numRewardSheets:
            rewardSumRef =  '=INDIRECT(\"' + sheets[iRewardSheet] + '!\"&ADDRESS(' + str(rewardSumRows[irow]) + ',' + str(rewardSumColumns[irow]) + ',4))'
            rewardSheet = sheets[iRewardSheet]
        else:
            rewardSumRef = ''
            rewardSheet = ''

        ws.append(['', '', '', tradeProfitSheet, tradeProfitSumRef, '', feeSheet, feeSumRef, '', rewardSheet, rewardSumRef])

    profitSumRow = ws.max_row + 2
    profitSumColumns = ['E', 'H', 'K']
    for profitSumColumn in profitSumColumns:
        ws[profitSumColumn + str(profitSumRow)] = '=SUM(' + profitSumColumn + str(
            firstProfitRow) + ':' + profitSumColumn + str(profitSumRow - 2) + ')'

    if translator:
        ws['A' + str(profitSumRow)] = trans('Profit')
    else:
        ws['A' + str(profitSumRow)] = 'Gewinn'
    profitSumSumColumn = 'B'
    cells = [col + str(profitSumRow) for col in profitSumColumns]
    ws[profitSumSumColumn + str(profitSumRow)] = ('=' + cells[0] + '+' + cells[1] + '+' + cells[2])

    # page setup
    pageSetup(ws, dateCols=[], gapCols=['C', 'F', 'I', 'L'], lastRow=profitSumRow, lastCol=6,
              setWidthCols=['A', 'B', 'D', 'G', 'J'], setWidthValue=[10, 23, 20, 20, 20], trans=trans)

    # def textLen(value):
    #     if value is None:
    #         return 2
    #     elif value == '':
    #         return 2
    #     elif isinstance(value, float):
    #         # return len(str(value))/2
    #         return 5
    #     elif isinstance(value, str) and '=' in value:
    #         return 5
    #     else:
    #         if str(value).islower():
    #             length = len(str(value)) * 0.9
    #         elif str(value).isupper():
    #             length = len(str(value)) * 1.2
    #         else:
    #             length = len(str(value)) + 2
    #         if length < 3:
    #             length = 3
    #         return length
    #
    # for ws in wb:
    #     for column_cells in ws.columns:
    #             length = max(textLen(cell.value) for cell in column_cells[1:])
    #             ws.column_dimensions[str(column_cells[1].column_letter)].width = length


    # save file
    # wb.save(path + '-' + str(datetime.datetime.now()).replace(' ', '_').replace(':', '_').replace('-', '_').replace('.',
    #                                                                                                                 '_') + '.xlsx')
    try:
        wb.save(path)
    except PermissionError as ex:
        localLogger.error("saving export failed: " + str(ex))


def pageSetup(ws, dateCols=[], gapCols=[], lastRow=None, lastCol=None, setWidthCols=[], setWidthValue=[], trans=None):
    # set width of columns
    for setWidthCol, setWidthVal in zip(setWidthCols, setWidthValue):
        ws.column_dimensions[setWidthCol].width = setWidthVal

    # set width of date columns
    for dateCol in dateCols:
        ws.column_dimensions[dateCol].width = 11

    # set width of gap columns
    gapFill = PatternFill(fill_type='solid',
                          start_color=GRAYCOLORCODE,
                          end_color=GRAYCOLORCODE)
    for gapCol in gapCols:
        ws.column_dimensions[gapCol].width = 3
        # set gap color
        for cell in ws[gapCol]:
            cell.fill = gapFill

    # add borders
    if lastRow and lastCol:
        for row in range(1, lastRow+1):
            for col in range(1, lastCol+1):
                ws.cell(row=row, column=col).border = thin_border

    # printer settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Header /Footer
    ws.oddHeader.left.text = "&[Date]"  # Date
    ws.oddHeader.left.size = 12
    ws.evenHeader.left.text = "&[Date]"  # Date
    ws.evenHeader.left.size = 12

    ws.oddHeader.center.text = "&A"  # TabName
    ws.oddHeader.center.size = 12
    ws.evenHeader.center.text = "&A"  # TabName
    ws.evenHeader.center.size = 12

    if trans:
        ws.oddFooter.right.text = trans("Page") + " &[Page] " + trans("of") + " &N"  # Page of Pages
        ws.evenFooter.right.text = trans("Page") + " &[Page] " + trans("of") + " &N"  # Page of Pages
    else:
        ws.oddFooter.right.text = "Seite &[Page] von &N"  # Page of Pages
        ws.evenFooter.right.text = "Seite &[Page] von &N"  # Page of Pages
    ws.oddFooter.right.size = 12
    ws.evenFooter.right.size = 12





